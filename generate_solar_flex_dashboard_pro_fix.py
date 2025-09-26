#!/usr/bin/env python3

# Robust Excel generator fallback (creates .xlsx if no template_with_vba.xlsm)
import os
import sys
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference

TEMPLATE = "template_with_vba.xlsm"
OUTPUT_XLSM = "solar_flex_dashboard_pro.xlsm"
OUTPUT_XLSX = "solar_flex_dashboard_pro.xlsx"
IMAGES_DIR = "images"
NUM_EXAMPLE_PROJECTS = 3

def ensure_dirs():
    if not os.path.exists(IMAGES_DIR):
        os.makedirs(IMAGES_DIR)

def create_basic_structure(wb):
    # Portada
    if "Portada" not in wb.sheetnames:
        ws = wb.create_sheet("Portada")
        ws["A1"] = "Solar Flex - Business Plan & Technical Dashboard (PRO)"
    # Proyectos base
    if "Proyectos" not in wb.sheetnames:
        ws_master = wb.create_sheet("Proyectos", 0)
    else:
        ws_master = wb["Proyectos"]
    headers = ["project_id", "Nombre Proyecto", "Cliente", "CUPS", "Potencia_kW", "Consumo_kWh", "Tipo Sistema", "Ubicación", "ImageFile"]
    for i, h in enumerate(headers, start=1):
        ws_master.cell(row=1, column=i, value=h)
    # Añadir ejemplos si no hay filas adicionales
    if ws_master.max_row == 1:
        for i in range(1, NUM_EXAMPLE_PROJECTS + 1):
            row = 1 + i
            ws_master.cell(row=row, column=1, value=f"project_{i}")
            ws_master.cell(row=row, column=2, value=f"Proyecto Ejemplo {i}")
            ws_master.cell(row=row, column=3, value=f"Cliente {i}")
            ws_master.cell(row=row, column=4, value=f"CUPS{i:04d}")
            ws_master.cell(row=row, column=5, value=50 * i)
            ws_master.cell(row=row, column=6, value=50000 * i)
            ws_master.cell(row=row, column=7, value="On-grid PV+ESS")
            ws_master.cell(row=row, column=8, value="Ciudad Ejemplo")
            ws_master.cell(row=row, column=9, value=f"project_{i}.png")
    return wb

def create_project_sheet(wb, project_row):
    pid = project_row[0].value
    name = project_row[1].value or pid
    sheet_name = str(name)[:31]
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    ws["A1"] = "Project ID"
    ws["B1"] = pid
    ws["A3"] = "Cliente"
    ws["B3"] = project_row[2].value or ""
    ws["A4"] = "CUPS"
    ws["B4"] = project_row[3].value or ""
    ws["A5"] = "Potencia instalada (kW)"
    ws["B5"] = project_row[4].value or 0
    ws["A6"] = "Consumo anual previsto (kWh)"
    ws["B6"] = project_row[5].value or 0
    ws["A7"] = "Tipo de sistema"
    ws["B7"] = project_row[6].value or ""
    ws["A8"] = "Ubicación"
    ws["B8"] = project_row[7].value or ""
    ws["A10"] = "Coste estimado (€)"
    ws["B10"] = "=B5*1200"
    ws["A11"] = "Ahorro anual (€)"
    ws["B11"] = "=B6*0.18"
    ws["A12"] = "CO2 evitado (ton)"
    ws["B12"] = "=B6*0.5/1000"
    img_file = project_row[8].value
    if img_file:
        img_path = os.path.join(IMAGES_DIR, img_file)
        if os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                img.width = 480
                img.height = 240
                ws.add_image(img, "D3")
            except Exception as e:
                print(f"[WARN] No se pudo insertar imagen {img_path}: {e}")
    return ws

def add_validations(ws_master):
    systems = ["On-grid PV+ESS", "Off-grid PV+ESS", "On/off-grid PV+ESS (ATS)", "Aislada", "Híbrido", "Generador", "Otros"]
    dv = DataValidation(type="list", formula1=f'"{", ".join(systems)}"', allow_blank=True)
    ws_master.add_data_validation(dv)
    try:
        dv.add("G2:G1000")
    except Exception:
        pass

def add_dashboard(wb):
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
    else:
        ws = wb.create_sheet("Dashboard", 1)
    ws["A1"] = "Dashboard Financiero y Estratégico"
    base = {"Portada","Proyectos","Dashboard","Supuestos","KPIs","Memoria Técnica","Esquemas"}
    project_sheets = [s for s in wb.sheetnames if s not in base]
    if project_sheets:
        sum_parts = [f"'{s}'!B10" for s in project_sheets]
        ws["A3"] = "Ventas proyectadas (€)"
        ws["B3"] = f"=SUM({','.join(sum_parts)})"
    row = 6
    for s in project_sheets:
        ws.cell(row=row, column=1, value=s)
        ws.cell(row=row, column=2, value=f"='{s}'!B10")
        row += 1
    try:
        if project_sheets:
            chart = BarChart()
            chart.title = "Coste estimado por proyecto"
            chart.y_axis.title = "€"
            data = Reference(ws, min_col=2, min_row=6, max_row=5+len(project_sheets))
            cats = Reference(ws, min_col=1, min_row=6, max_row=5+len(project_sheets))
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            ws.add_chart(chart, "E6")
    except Exception as e:
        print(f"[WARN] No se pudo crear gráfico: {e}")

def run_with_template():
    print("Intentando cargar plantilla con VBA (si existe)...")
    wb = load_workbook(TEMPLATE, keep_vba=True)
    create_basic_structure(wb)
    ws_master = wb["Proyectos"]
    add_validations(ws_master)
    for row in ws_master.iter_rows(min_row=2, values_only=False):
        if row[0].value:
            create_project_sheet(wb, row)
    if "Supuestos" not in wb.sheetnames:
        ws_sup = wb.create_sheet("Supuestos")
        ws_sup["A1"] = "Años del plan"
        ws_sup["B1"] = "2025-2030"
    add_dashboard(wb)
    wb.save(OUTPUT_XLSM)
    print(f"[OK] Generado: {OUTPUT_XLSM}")

def run_without_template():
    print("Plantilla con VBA no encontrada. Generando versión .xlsx de prueba (sin macros)...")
    wb = Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        std = wb["Sheet"]
        wb.remove(std)
    create_basic_structure(wb)
    ws_master = wb["Proyectos"]
    add_validations(ws_master)
    for row in ws_master.iter_rows(min_row=2, values_only=False):
        if row[0].value:
            create_project_sheet(wb, row)
    if "Supuestos" not in wb.sheetnames:
        ws_sup = wb.create_sheet("Supuestos")
        ws_sup["A1"] = "Años del plan"
        ws_sup["B1"] = "2025-2030"
    add_dashboard(wb)
    wb.save(OUTPUT_XLSX)
    print(f"[OK] Generado (sin macros): {OUTPUT_XLSX}")

def main():
    ensure_dirs()
    try:
        if os.path.exists(TEMPLATE):
            run_with_template()
        else:
            run_without_template()
    except Exception as e:
        print("[ERROR] Ocurrió una excepción:")
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
